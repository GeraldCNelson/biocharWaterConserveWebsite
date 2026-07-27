#!/usr/bin/env python3
"""
update_master_workbook_snapshot.py
Copy and validate a snapshot of the synchronized master project workbook.

Purpose
-------
Copy the authoritative master workbook from its local OneDrive synchronized
location into the project's canonical raw-data location.

The synchronized source, destination, view-only source URL, update method, and
required worksheets are registered in
``biochar_app.config.data_sources.BIOCHAR_MASTER_SOURCE``.

Safety
------
- The synchronized OneDrive workbook is opened read-only and never modified.
- The source is copied to a temporary file before validation.
- The current project snapshot remains untouched if copying or validation fails.
- The destination is replaced atomically only after successful validation.
- An audit JSON file records paths, worksheet checks, sizes, and SHA-256 hashes.

Run from the repository root
----------------------------

Validate without replacing the project snapshot::

    python biochar_app/scripts/management/update_master_workbook_snapshot.py \
        --validate-only

Validate and update the project snapshot::

    python biochar_app/scripts/management/update_master_workbook_snapshot.py

Introduced
----------
2026-07-25

Maintainer
----------
Biochar Water Conservation project (Gerald Nelson).
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from biochar_app.config.data_sources import BIOCHAR_MASTER_SOURCE


READ_CHUNK_SIZE = 1024 * 1024


def require_onedrive_desktop_app() -> None:
    """
    Stop when the macOS OneDrive synchronization application is not running.

    This check confirms only that the desktop application is running. Workbook
    existence, local availability, validity, and copy integrity are checked
    separately by ``update_snapshot``.
    """
    if sys.platform != "darwin":
        raise RuntimeError(
            "Automatic master-workbook refresh requires the Microsoft "
            "OneDrive desktop application on macOS. Rerun ETL with "
            "--skip-master-workbook-refresh when using a previously "
            "installed repository snapshot."
        )

    try:
        result = subprocess.run(
            ["pgrep", "-x", "OneDrive"],
            capture_output=True,
            text=True,
            check=False,
        )
    except FileNotFoundError as error:
        raise RuntimeError(
            "Unable to check whether Microsoft OneDrive is running because "
            "the macOS 'pgrep' command was not found."
        ) from error

    if result.returncode != 0:
        raise RuntimeError(
            "Microsoft OneDrive is not running. Start OneDrive, wait for "
            "synchronization to finish, and rerun the command.\n"
            "Expected synchronized workbook:\n"
            f"{BIOCHAR_MASTER_SOURCE.synced_source_path}"
        )


def sha256_file(path: Path) -> str:
    """Return the SHA-256 digest of a file."""
    digest = hashlib.sha256()

    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(READ_CHUNK_SIZE), b""):
            digest.update(chunk)

    return digest.hexdigest()


def normalize_sheet_name(name: str) -> str:
    """
    Normalize harmless worksheet-name whitespace for validation.

    Excel may contain leading, trailing, or repeated spaces that are difficult
    to see in its worksheet tabs. Normalization does not rename worksheets.
    """
    return " ".join(str(name).split()).casefold()


def validate_workbook(
    path: Path,
    required_sheets: tuple[str, ...],
) -> dict[str, Any]:
    """
    Validate that ``path`` is an XLSX workbook with required worksheets.

    Required worksheet names are compared case-insensitively after whitespace
    normalization. Actual worksheet names are retained in the audit report.
    """
    if not path.exists():
        raise FileNotFoundError(f"Workbook does not exist: {path}")

    if not path.is_file():
        raise ValueError(f"Workbook path is not a file: {path}")

    size_bytes = path.stat().st_size

    if size_bytes == 0:
        raise ValueError(f"Workbook is empty: {path}")

    if not zipfile.is_zipfile(path):
        raise ValueError(
            f"File is not a valid XLSX/ZIP workbook: {path}"
        )

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=False,
    )

    try:
        actual_sheet_names = tuple(workbook.sheetnames)
    finally:
        workbook.close()

    normalized_actual: dict[str, list[str]] = {}

    for actual_name in actual_sheet_names:
        normalized = normalize_sheet_name(actual_name)
        normalized_actual.setdefault(normalized, []).append(actual_name)

    ambiguous_names = {
        normalized: names
        for normalized, names in normalized_actual.items()
        if len(names) > 1
    }

    if ambiguous_names:
        raise ValueError(
            "Workbook contains worksheet names that become ambiguous after "
            f"normalization: {ambiguous_names}"
        )

    required_matches: dict[str, str] = {}
    missing_required_sheets: list[str] = []

    for required_name in required_sheets:
        normalized_required = normalize_sheet_name(required_name)
        matches = normalized_actual.get(normalized_required, [])

        if not matches:
            missing_required_sheets.append(required_name)
        else:
            required_matches[required_name] = matches[0]

    if missing_required_sheets:
        raise ValueError(
            "Workbook is missing required worksheets: "
            f"{missing_required_sheets}. "
            f"Available worksheets: {list(actual_sheet_names)}"
        )

    return {
        "size_bytes": size_bytes,
        "sheet_count": len(actual_sheet_names),
        "sheet_names": list(actual_sheet_names),
        "required_sheets": list(required_sheets),
        "required_sheet_matches": required_matches,
        "missing_required_sheets": [],
    }


def write_audit(
    audit_path: Path,
    audit: dict[str, Any],
) -> None:
    """Write the audit JSON atomically."""
    audit_path.parent.mkdir(parents=True, exist_ok=True)

    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        prefix=f".{audit_path.stem}-",
        suffix=".json",
        dir=audit_path.parent,
        delete=False,
    ) as temporary_file:
        temporary_audit_path = Path(temporary_file.name)
        json.dump(
            audit,
            temporary_file,
            indent=2,
            sort_keys=True,
        )
        temporary_file.write("\n")

    try:
        os.replace(temporary_audit_path, audit_path)
    finally:
        if temporary_audit_path.exists():
            temporary_audit_path.unlink()


def update_snapshot(
    *,
    source: Path,
    destination: Path,
    required_sheets: tuple[str, ...],
    audit_path: Path,
    validate_only: bool,
) -> dict[str, Any]:
    """
    Copy, validate, and optionally install the master-workbook snapshot.
    """
    source = source.expanduser()
    destination = destination.expanduser()
    audit_path = audit_path.expanduser()

    if not source.exists():
        raise FileNotFoundError(
            "The synchronized OneDrive workbook was not found:\n"
            f"{source}\n"
            "Confirm that OneDrive is synchronized and the workbook is "
            "available locally."
        )

    if not source.is_file():
        raise ValueError(
            f"The synchronized workbook path is not a file: {source}"
        )

    destination.parent.mkdir(parents=True, exist_ok=True)

    previous_destination_sha256 = (
        sha256_file(destination)
        if destination.exists()
        else None
    )

    temporary_snapshot: Path | None = None

    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            prefix=f".{destination.stem}-",
            suffix=".xlsx",
            dir=destination.parent,
            delete=False,
        ) as temporary_file:
            temporary_snapshot = Path(temporary_file.name)

        shutil.copy2(source, temporary_snapshot)

        validation = validate_workbook(
            temporary_snapshot,
            required_sheets,
        )

        source_sha256 = sha256_file(source)
        copied_sha256 = sha256_file(temporary_snapshot)

        if source_sha256 != copied_sha256:
            raise ValueError(
                "The temporary copy does not match the synchronized source. "
                f"Source SHA-256: {source_sha256}; "
                f"copy SHA-256: {copied_sha256}"
            )

        changed = copied_sha256 != previous_destination_sha256

        audit: dict[str, Any] = {
            "source_key": BIOCHAR_MASTER_SOURCE.key,
            "provider": BIOCHAR_MASTER_SOURCE.provider,
            "description": BIOCHAR_MASTER_SOURCE.description,
            "source_url": BIOCHAR_MASTER_SOURCE.source_url,
            "update_method": BIOCHAR_MASTER_SOURCE.update_method,
            "retrieved_at_utc": datetime.now(timezone.utc).isoformat(),
            "synced_source_path": str(source),
            "destination_path": str(destination),
            "validate_only": validate_only,
            "installed": False,
            "changed": changed,
            "source_sha256": source_sha256,
            "copied_sha256": copied_sha256,
            "previous_destination_sha256": previous_destination_sha256,
            **validation,
        }

        if validate_only:
            audit["result"] = "validated_not_installed"
        else:
            os.replace(temporary_snapshot, destination)
            temporary_snapshot = None

            audit["installed"] = True
            audit["installed_sha256"] = sha256_file(destination)
            audit["result"] = (
                "installed_changed_snapshot"
                if changed
                else "installed_identical_snapshot"
            )

        write_audit(audit_path, audit)
        return audit

    finally:
        if (
            temporary_snapshot is not None
            and temporary_snapshot.exists()
        ):
            temporary_snapshot.unlink()


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""
    default_audit_path = (
        BIOCHAR_MASTER_SOURCE.local_path.with_suffix(
            ".snapshot.json"
        )
    )

    parser = argparse.ArgumentParser(
        description=(
            "Copy and validate the synchronized master workbook."
        )
    )

    parser.add_argument(
        "--source",
        type=Path,
        default=BIOCHAR_MASTER_SOURCE.synced_source_path,
        help=(
            "Synchronized OneDrive workbook. "
            f"Default: {BIOCHAR_MASTER_SOURCE.synced_source_path}"
        ),
    )

    parser.add_argument(
        "--destination",
        type=Path,
        default=BIOCHAR_MASTER_SOURCE.local_path,
        help=(
            "Validated project snapshot. "
            f"Default: {BIOCHAR_MASTER_SOURCE.local_path}"
        ),
    )

    parser.add_argument(
        "--audit-json",
        type=Path,
        default=default_audit_path,
        help=(
            "Audit JSON output. "
            f"Default: {default_audit_path}"
        ),
    )

    parser.add_argument(
        "--validate-only",
        action="store_true",
        help=(
            "Copy and validate without replacing the project snapshot."
        ),
    )

    return parser.parse_args()


def main() -> None:
    """Run the master-workbook snapshot update."""
    args = parse_args()

    audit = update_snapshot(
        source=args.source,
        destination=args.destination,
        required_sheets=(
            BIOCHAR_MASTER_SOURCE.required_sheets
        ),
        audit_path=args.audit_json,
        validate_only=args.validate_only,
    )

    print(json.dumps(audit, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()