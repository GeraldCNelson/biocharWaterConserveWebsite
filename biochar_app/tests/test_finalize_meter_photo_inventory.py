from __future__ import annotations

import csv
import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


SCRIPT = (
    Path(__file__).parents[1]
    / "scripts/management/finalize_meter_photo_inventory.py"
)
SPEC = importlib.util.spec_from_file_location("finalize_meter_photo_inventory", SCRIPT)
MODULE = importlib.util.module_from_spec(SPEC)
assert SPEC and SPEC.loader
sys.modules[SPEC.name] = MODULE
SPEC.loader.exec_module(MODULE)


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def make_main(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Meter Review"
    sheet.append(["Filename", "SHA-256", "Manual Reading", "Status", "Notes"])
    sheet.append(["UUID.heic", "sha-original", "162023", "readable", ""])
    sheet.append(["UUID_4_5005_c.jpeg", "sha-small", "", "unreadable", ""])
    sheet.append(["IMG_7515.JPEG", "sha-img", "", "unreadable", ""])
    workbook.save(path)


def make_followup(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, values in enumerate(
        [
            ("UUID_4_5005_c.jpeg", "sha-small", "182023"),
            ("IMG_7515.JPEG", "sha-img", "182023"),
        ],
        start=1,
    ):
        sheet = workbook.create_sheet(f"Photo {index:02d}")
        sheet["B5"], sheet["B6"], sheet["B11"], sheet["B12"] = (
            values[0], values[1], values[2], "readable"
        )
    workbook.save(path)


class FinalizeInventoryTests(unittest.TestCase):
    def test_precedence_and_unique_selection(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            inventory = tmp_path / "photo_inventory.csv"
            main = tmp_path / "main.xlsx"
            followup = tmp_path / "followup.xlsx"
            corrections = tmp_path / "corrections.csv"
            output = tmp_path / "photo_inventory_unique.csv"
            manifest = tmp_path / "manifest.csv"
            audit = tmp_path / "audit.json"

            rows = [
                {
                    "filename": "UUID.heic", "sha256": "sha-original",
                    "effective_datetime": "2023-05-23 19:19:29", "meter_reading": "",
                    "review_status": "", "image_width": "4032", "image_height": "3024",
                    "file_size_bytes": "2000000",
                },
                {
                    "filename": "UUID_4_5005_c.jpeg", "sha256": "sha-small",
                    "effective_datetime": "2023-05-23 19:19:29", "meter_reading": "",
                    "review_status": "", "image_width": "360", "image_height": "480",
                    "file_size_bytes": "70000",
                },
                {
                    "filename": "IMG_7515.JPEG", "sha256": "sha-img",
                    "effective_datetime": "2023-05-23 19:19:29", "meter_reading": "",
                    "review_status": "", "image_width": "1536", "image_height": "2048",
                    "file_size_bytes": "700000",
                },
            ]
            write_csv(inventory, rows)
            make_main(main)
            make_followup(followup)
            write_csv(
                corrections,
                [{
                    "sha256": "sha-img", "filename": "IMG_7515.JPEG",
                    "corrected_reading": "162023", "reason": "confirmed",
                }],
            )

            result = MODULE.finalize_inventory(
                inventory, main, followup, corrections, output, manifest, audit
            )

            _, selected = MODULE.read_csv_rows(output)
            self.assertEqual(len(selected), 1)
            self.assertEqual(selected[0]["filename"], "UUID.heic")
            self.assertEqual(selected[0]["meter_reading"], "162023")
            self.assertEqual(
                result["suppressed_followup_conflicts"][0]["selected_reading"],
                "162023",
            )
            self.assertEqual(
                json.loads(audit.read_text())["explicit_correction_count"],
                1,
            )

    def test_conflicting_readable_main_family_is_not_propagated(self) -> None:
        rows = [
            MODULE.ReviewReading("a", "UUID.heic", "162023", "readable", "", "main"),
            MODULE.ReviewReading("b", "UUID_4_5005_c.jpeg", "182023", "readable", "", "main"),
        ]
        consensus, ambiguous = MODULE.main_family_consensus(rows)
        self.assertNotIn("uuid", consensus)
        self.assertEqual(ambiguous["uuid"], ["162023", "182023"])


if __name__ == "__main__":
    unittest.main()
