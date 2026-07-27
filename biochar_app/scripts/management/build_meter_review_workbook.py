#!/usr/bin/env python3
"""
Build an Excel workbook for manual transcription of irrigation meter photos.

The workbook embeds a cropped meter-head thumbnail beside an editable six-digit
manual-reading field. It reads canonical photos and effective timestamps from
the meter-photo inventory. Duplicate management belongs to the inventory; this
workbook includes each inventory photo once.

Default input
-------------
biochar_app/data-processed/management/irrigation/photos/originals/
biochar_app/data-processed/management/irrigation/photos/photo_inventory.csv

Default output
--------------
biochar_app/data-processed/management/irrigation/photos/
    meter_photo_review.xlsx

Run from the repository root:

    python biochar_app/scripts/management/build_meter_review_workbook.py

Optional examples:

    python biochar_app/scripts/management/build_meter_review_workbook.py \
      --max-images 25

    python biochar_app/scripts/management/build_meter_review_workbook.py \
      --output-xlsx /tmp/meter_review.xlsx

Dependencies
------------
    pip install openpyxl pillow pillow-heif opencv-python pandas
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import os
import re
import tempfile
import warnings
from pathlib import Path
import cv2
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image, ImageOps

from biochar_app.config.paths import IRRIGATION_DIR

try:
    from pillow_heif import register_heif_opener
except ImportError:  # pragma: no cover - handled at runtime
    register_heif_opener = None


DEFAULT_PHOTOS_ROOT = IRRIGATION_DIR / "photos"
DEFAULT_PHOTO_DIR = DEFAULT_PHOTOS_ROOT / "originals"
DEFAULT_METADATA_CSV = DEFAULT_PHOTOS_ROOT / "photo_inventory.csv"
DEFAULT_OUTPUT_XLSX = DEFAULT_PHOTOS_ROOT / "meter_photo_review.xlsx"
DEFAULT_CROP_DIR = DEFAULT_PHOTOS_ROOT / "meter_review_workbook_crops"
DEFAULT_CORRECTIONS_CSV = (
    DEFAULT_PHOTOS_ROOT / "meter_photo_reading_corrections.csv"
)

SUPPORTED_EXTENSIONS = {
    ".jpg",
    ".jpeg",
    ".png",
    ".heic",
    ".heif",
}

HEIC_EXTENSIONS = {".heic", ".heif"}

STATUS_VALUES = (
    "readable",
    "uncertain",
    "unreadable",
    "not_meter_reading",
    "duplicate",
)

THUMBNAIL_WIDTH_PX = 360
THUMBNAIL_HEIGHT_PX = 260
ROW_HEIGHT_POINTS = 205


# ---------------------------------------------------------------------------
# CLI and metadata handling
# ---------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build an Excel workbook with cropped meter photos beside manual "
            "reading fields."
        )
    )

    parser.add_argument(
        "--photo-dir",
        type=Path,
        default=DEFAULT_PHOTO_DIR,
        help=f"Photo directory. Default: {DEFAULT_PHOTO_DIR}",
    )

    parser.add_argument(
        "--metadata-csv",
        type=Path,
        default=DEFAULT_METADATA_CSV,
        help=f"Photo metadata CSV. Default: {DEFAULT_METADATA_CSV}",
    )

    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=DEFAULT_OUTPUT_XLSX,
        help=f"Output workbook. Default: {DEFAULT_OUTPUT_XLSX}",
    )

    parser.add_argument(
        "--corrections-csv",
        type=Path,
        default=DEFAULT_CORRECTIONS_CSV,
        help=(
            "Optional SHA-256-keyed manual-reading corrections ledger. "
            f"Default: {DEFAULT_CORRECTIONS_CSV}"
        ),
    )

    parser.add_argument(
        "--crop-dir",
        type=Path,
        default=DEFAULT_CROP_DIR,
        help=f"Generated meter-head crop directory. Default: {DEFAULT_CROP_DIR}",
    )

    parser.add_argument(
        "--max-images",
        type=int,
        default=None,
        help="Optional maximum number of inventory photos to include.",
    )

    parser.add_argument(
        "--keep-crops",
        action="store_true",
        help=(
            "Keep generated meter-head JPEG crops after the workbook is saved. "
            "The workbook embeds the images, so keeping them is optional."
        ),
    )

    parser.add_argument(
        "--allow-reading-loss",
        action="store_true",
        help=(
            "Allow replacement when the new workbook contains fewer reviewed "
            "readings than the existing output. Intended only for deliberate "
            "recovery work."
        ),
    )

    return parser.parse_args()


def find_column(
    df: pd.DataFrame,
    candidates: tuple[str, ...],
) -> str | None:
    normalized = {
        str(column).strip().lower(): str(column)
        for column in df.columns
    }

    for candidate in candidates:
        found = normalized.get(candidate.lower())
        if found is not None:
            return found

    return None


def normalize_filename(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    return Path(str(value).strip()).name


def clean_text(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


def valid_meter_reading(value: object) -> bool:
    return bool(re.fullmatch(r"\d{6}", clean_text(value)))


def load_corrections(corrections_csv: Path | None) -> dict[str, dict[str, str]]:
    if corrections_csv is None or not corrections_csv.exists():
        return {}

    with corrections_csv.open(
        newline="",
        encoding="utf-8-sig",
    ) as handle:
        reader = csv.DictReader(handle)
        required = {
            "sha256",
            "filename",
            "corrected_reading",
            "reason",
        }
        missing = required.difference(reader.fieldnames or [])
        if missing:
            raise KeyError(
                "Corrections ledger is missing columns: "
                f"{sorted(missing)}"
            )

        corrections: dict[str, dict[str, str]] = {}
        for row in reader:
            sha256 = clean_text(row["sha256"]).lower()
            reading = clean_text(row["corrected_reading"])
            if not re.fullmatch(r"[0-9a-f]{64}", sha256):
                raise ValueError(
                    f"Invalid SHA-256 in corrections ledger: {sha256!r}"
                )
            if not valid_meter_reading(reading):
                raise ValueError(
                    "Invalid six-digit correction for "
                    f"{clean_text(row['filename'])!r}: {reading!r}"
                )
            corrections[sha256] = {
                "filename": clean_text(row["filename"]),
                "meter_reading": reading,
                "review_status": "readable",
                "notes": clean_text(row["reason"]),
            }
        return corrections


def apply_corrections(
    metadata: pd.DataFrame,
    corrections: dict[str, dict[str, str]],
) -> pd.DataFrame:
    out = metadata.copy()
    if not corrections:
        return out
    if "sha256" not in out.columns:
        raise KeyError(
            "Metadata CSV must contain sha256 when corrections are supplied."
        )

    normalized_sha = out["sha256"].map(clean_text).str.lower()
    inventory_hashes = set(normalized_sha)
    unmatched = sorted(set(corrections) - inventory_hashes)
    if unmatched:
        warnings.warn(
            "Corrections ledger contains SHA-256 values not present in this "
            "metadata subset. They may belong to excluded duplicate photos: "
            f"{unmatched}",
            stacklevel=2,
        )

    for sha256, correction in corrections.items():
        mask = normalized_sha.eq(sha256)
        out.loc[mask, "meter_reading"] = correction["meter_reading"]
        out.loc[mask, "review_status"] = correction["review_status"]
        out.loc[mask, "notes"] = correction["notes"]

    return out


def load_metadata(
    metadata_csv: Path,
    photo_dir: Path,
) -> pd.DataFrame:
    if not metadata_csv.exists():
        raise FileNotFoundError(
            f"Metadata CSV does not exist: {metadata_csv}"
        )

    df = pd.read_csv(
        metadata_csv,
        dtype=str,
        keep_default_na=False,
    )

    filename_col = find_column(
        df,
        (
            "filename",
            "file_name",
            "sourcefile",
            "source_file",
        ),
    )

    datetime_col = find_column(
        df,
        (
            "effective_datetime",
            "manual_datetime",
            "selected_datetime",
            "datetime_original",
            "photo_datetime",
            "create_date",
            "timestamp",
        ),
    )

    if filename_col is None:
        raise KeyError(
            "Could not find a filename column in the metadata CSV."
        )

    if datetime_col is None:
        raise KeyError(
            "Could not find a photo datetime column in the metadata CSV."
        )

    out = df.copy()
    out["filename"] = out[filename_col].map(normalize_filename)
    out["photo_datetime"] = pd.to_datetime(
        out[datetime_col],
        errors="coerce",
    )
    out["extension"] = out["filename"].map(
        lambda value: Path(value).suffix.lower()
    )
    out["image_path"] = out["filename"].map(
        lambda value: photo_dir / value
    )

    out = out.loc[
        out["filename"].ne("")
        & out["extension"].isin(SUPPORTED_EXTENSIONS)
        & out["image_path"].map(Path.exists)
    ].copy()

    return out


def select_inventory_photos(
    metadata: pd.DataFrame,
) -> pd.DataFrame:
    """
    Return every canonical inventory photo, ordered by timestamp and filename.

    Do not group by timestamp: two distinct photographs may legitimately have
    the same timestamp. Exact-file duplicate management belongs to the photo
    inventory and its cleanup workflow.
    """
    out = metadata.copy()
    duplicate_count_col = find_column(
        out,
        ("exact_duplicate_count", "duplicate_count"),
    )
    if duplicate_count_col is None:
        out["duplicate_count"] = 1
    else:
        out["duplicate_count"] = (
            pd.to_numeric(out[duplicate_count_col], errors="coerce")
            .fillna(1)
            .clip(lower=1)
            .astype(int)
        )

    return out.sort_values(
        ["photo_datetime", "filename"],
        na_position="last",
        kind="stable",
    ).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Image loading and meter-head cropping
# ---------------------------------------------------------------------------


def register_heic_support() -> bool:
    if register_heif_opener is None:
        return False

    register_heif_opener()
    return True


def read_image(
    image_path: Path,
    heic_available: bool,
) -> np.ndarray:
    if (
        image_path.suffix.lower() in HEIC_EXTENSIONS
        and not heic_available
    ):
        raise RuntimeError(
            "HEIC support is required. Install it with: pip install pillow-heif"
        )

    with Image.open(image_path) as image:
        image = ImageOps.exif_transpose(image).convert("RGB")
        rgb = np.asarray(image)

    return cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)


def resize_for_detection(
    image: np.ndarray,
    max_dimension: int = 1800,
) -> tuple[np.ndarray, float]:
    height, width = image.shape[:2]
    largest = max(height, width)

    if largest <= max_dimension:
        return image.copy(), 1.0

    scale = max_dimension / largest
    resized = cv2.resize(
        image,
        None,
        fx=scale,
        fy=scale,
        interpolation=cv2.INTER_AREA,
    )

    return resized, scale


def detect_meter_face(
    image: np.ndarray,
) -> tuple[int, int, int] | None:
    """
    Detect the circular meter face using a Hough-circle search.
    """
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (9, 9), 2)

    height, width = gray.shape
    min_dimension = min(height, width)

    circles = cv2.HoughCircles(
        gray,
        cv2.HOUGH_GRADIENT,
        dp=1.2,
        minDist=min_dimension * 0.25,
        param1=100,
        param2=45,
        minRadius=int(min_dimension * 0.12),
        maxRadius=int(min_dimension * 0.48),
    )

    if circles is None:
        return None

    candidates = np.round(circles[0]).astype(int)
    image_center_x = width / 2.0
    image_center_y = height / 2.0

    def candidate_score(circle: np.ndarray) -> float:
        center_x, center_y, radius = circle
        center_distance = np.hypot(
            center_x - image_center_x,
            center_y - image_center_y,
        )
        return float(radius) - 0.35 * float(center_distance)

    best = max(candidates, key=candidate_score)
    center_x, center_y, radius = map(int, best)

    return center_x, center_y, radius


def fixed_meter_head_crop(image: np.ndarray) -> np.ndarray:
    """
    Broad fallback crop for photographs where circle detection fails.
    """
    height, width = image.shape[:2]

    x1 = int(width * 0.12)
    x2 = int(width * 0.88)
    y1 = int(height * 0.18)
    y2 = int(height * 0.88)

    return image[y1:y2, x1:x2].copy()


def crop_meter_head(
    image: np.ndarray,
) -> tuple[np.ndarray, str]:
    """
    Crop the complete meter head, retaining enough dial context for a human.
    """
    meter_face = detect_meter_face(image)

    if meter_face is None:
        return fixed_meter_head_crop(image), "fixed_fallback"

    center_x, center_y, radius = meter_face
    image_height, image_width = image.shape[:2]

    # Include the circular meter face plus a modest rim. This is intentionally
    # broader than the OCR crop so the six-wheel reading remains easy to verify.
    horizontal_radius = 1.18 * radius
    upper_radius = 1.12 * radius
    lower_radius = 1.28 * radius

    x1 = max(0, int(center_x - horizontal_radius))
    x2 = min(image_width, int(center_x + horizontal_radius))
    y1 = max(0, int(center_y - upper_radius))
    y2 = min(image_height, int(center_y + lower_radius))

    crop = image[y1:y2, x1:x2].copy()

    if crop.size == 0:
        return fixed_meter_head_crop(image), "fixed_fallback"

    return crop, "circle_meter_head"


def save_workbook_crop(
    image_path: Path,
    crop_dir: Path,
    heic_available: bool,
) -> tuple[Path, str]:
    image = read_image(
        image_path,
        heic_available=heic_available,
    )

    resized, _ = resize_for_detection(image)
    crop, method = crop_meter_head(resized)

    if crop.size == 0:
        raise ValueError("Meter-head crop is empty")

    crop_dir.mkdir(parents=True, exist_ok=True)

    safe_stem = re.sub(
        r"[^A-Za-z0-9._-]+",
        "_",
        image_path.stem,
    )
    path_hash = hashlib.sha1(
        str(image_path.resolve()).encode("utf-8")
    ).hexdigest()[:10]
    crop_path = crop_dir / f"{safe_stem}_{path_hash}_meter_head.jpg"

    written = cv2.imwrite(
        str(crop_path),
        crop,
        [int(cv2.IMWRITE_JPEG_QUALITY), 90],
    )
    if not written or not crop_path.is_file():
        raise OSError(f"Could not write workbook crop: {crop_path}")

    return crop_path, method


# ---------------------------------------------------------------------------
# Workbook construction
# ---------------------------------------------------------------------------


def fit_thumbnail(
    image_path: Path,
    max_width: int,
    max_height: int,
) -> tuple[int, int]:
    with Image.open(image_path) as image:
        width, height = image.size

    scale = min(
        max_width / width,
        max_height / height,
        1.0,
    )

    return (
        max(1, int(width * scale)),
        max(1, int(height * scale)),
    )


def add_instructions_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Instructions")

    rows = [
        ["Meter Photo Review Workbook"],
        [""],
        [
            "Enter the six-digit mechanical meter reading in the Manual Reading "
            "column beside each photo."
        ],
        [
            "Use Status = readable when the value is clear; uncertain when one "
            "or more digits need a second look; unreadable when the display cannot "
            "be resolved; not_meter_reading for unrelated diagnostic photos."
        ],
        [
            "The meter face reports GALLONS × 100. Enter the six displayed digits "
            "only; do not multiply by 100 in the workbook."
        ],
        [
            "The Filename cell hyperlinks to the canonical original photo. The "
            "duplicate count comes from the photo inventory and should normally be 1."
        ],
    ]

    for row in rows:
        sheet.append(row)

    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("A1:F1")

    for row_number in range(3, 7):
        sheet.cell(row=row_number, column=1).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

    sheet.column_dimensions["A"].width = 105
    sheet.sheet_view.showGridLines = False


def build_workbook(
    selected_photos: pd.DataFrame,
    output_xlsx: Path,
    crop_dir: Path,
    heic_available: bool,
    keep_crops: bool,
) -> None:
    workbook = Workbook()
    review = workbook.active
    review.title = "Meter Review"
    add_instructions_sheet(workbook)

    headers = [
        "Photo Date/Time",
        "Photo",
        "Manual Reading",
        "Status",
        "Notes",
        "Filename",
        "SHA-256",
        "Crop Method",
        "Duplicate Count",
    ]

    review.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )

    for cell in review[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = border

    review.row_dimensions[1].height = 34

    status_validation = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUS_VALUES) + '"',
        allow_blank=True,
    )
    status_validation.error = "Choose a value from the status list."
    status_validation.errorTitle = "Invalid status"
    review.add_data_validation(status_validation)

    # Accept exactly six numeric digits, while allowing blank cells during review.
    reading_validation = DataValidation(
        type="custom",
        formula1=(
            '=OR(C2="",AND(LEN(C2)=6,ISNUMBER(--C2)))'
        ),
        allow_blank=True,
    )
    reading_validation.error = (
        "Enter exactly six digits, or leave the cell blank."
    )
    reading_validation.errorTitle = "Invalid meter reading"
    review.add_data_validation(reading_validation)

    image_failures = 0
    generated_crops: list[Path] = []

    for index, row in selected_photos.iterrows():
        excel_row = index + 2
        photo_datetime = row.get("photo_datetime")
        image_path = Path(str(row["image_path"]))

        crop_path: Path | None = None
        crop_method = ""
        crop_error = ""

        try:
            crop_path, crop_method = save_workbook_crop(
                image_path=image_path,
                crop_dir=crop_dir,
                heic_available=heic_available,
            )
            generated_crops.append(crop_path)
        except Exception as exc:  # keep workbook build moving
            image_failures += 1
            crop_error = f"{type(exc).__name__}: {exc}"
            crop_method = "crop_failed"

        values = [
            (
                photo_datetime.to_pydatetime()
                if pd.notna(photo_datetime)
                else None
            ),
            "",
            clean_text(row.get("meter_reading")),
            clean_text(row.get("review_status")),
            clean_text(row.get("notes")) or crop_error,
            image_path.name,
            clean_text(row.get("sha256")),
            crop_method,
            int(row.get("duplicate_count", 1)),
        ]

        review.append(values)

        review.row_dimensions[excel_row].height = ROW_HEIGHT_POINTS

        for column_index in range(1, len(headers) + 1):
            cell = review.cell(
                row=excel_row,
                column=column_index,
            )
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )
            cell.border = border

        review.cell(excel_row, 1).number_format = "yyyy-mm-dd hh:mm:ss"
        review.cell(excel_row, 3).number_format = "@"
        review.cell(excel_row, 3).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        review.cell(excel_row, 3).font = Font(size=14, bold=True)

        filename_cell = review.cell(excel_row, 6)
        filename_cell.hyperlink = image_path.resolve().as_uri()
        filename_cell.style = "Hyperlink"

        if crop_path is not None and crop_path.exists():
            thumbnail_width, thumbnail_height = fit_thumbnail(
                crop_path,
                max_width=THUMBNAIL_WIDTH_PX,
                max_height=THUMBNAIL_HEIGHT_PX,
            )

            excel_image = XLImage(str(crop_path))
            excel_image.width = thumbnail_width
            excel_image.height = thumbnail_height
            excel_image.anchor = f"B{excel_row}"
            review.add_image(excel_image)

    last_row = len(selected_photos) + 1

    reading_validation.add(f"C2:C{last_row}")
    status_validation.add(f"D2:D{last_row}")

    # Visual cues for rows still needing work.
    review.conditional_formatting.add(
        f"A2:I{last_row}",
        FormulaRule(
            formula=["$C2=\"\""],
            fill=PatternFill("solid", fgColor="FFF2CC"),
        ),
    )
    review.conditional_formatting.add(
        f"A2:I{last_row}",
        FormulaRule(
            formula=['$D2="uncertain"'],
            fill=PatternFill("solid", fgColor="FCE4D6"),
        ),
    )
    review.conditional_formatting.add(
        f"A2:I{last_row}",
        FormulaRule(
            formula=['$D2="readable"'],
            fill=PatternFill("solid", fgColor="E2F0D9"),
        ),
    )

    review.freeze_panes = "A2"
    review.auto_filter.ref = f"A1:I{last_row}"
    review.sheet_view.showGridLines = False

    # Photo and Manual Reading are deliberately adjacent.
    column_widths = {
        "A": 22,
        "B": 52,
        "C": 18,
        "D": 22,
        "E": 34,
        "F": 48,
        "G": 68,
        "H": 20,
        "I": 15,
    }

    for column, width in column_widths.items():
        review.column_dimensions[column].width = width

    review.sheet_properties.pageSetUpPr.fitToPage = True
    review.page_setup.fitToWidth = 1
    review.page_setup.fitToHeight = 0
    review.print_title_rows = "1:1"

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    temporary_output: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{output_xlsx.stem}-",
            suffix=".xlsx",
            dir=output_xlsx.parent,
            delete=False,
        ) as handle:
            temporary_output = Path(handle.name)
        workbook.save(temporary_output)
        os.replace(temporary_output, output_xlsx)
        temporary_output = None
    finally:
        if temporary_output is not None:
            temporary_output.unlink(missing_ok=True)

    if not keep_crops:
        for crop_path in generated_crops:
            crop_path.unlink(missing_ok=True)
        try:
            crop_dir.rmdir()
        except OSError:
            # Preserve a non-empty directory containing files from another run.
            pass

    print(f"Inventory photo rows: {len(selected_photos)}")
    print(f"Crop failures: {image_failures}")
    print(f"Wrote workbook: {output_xlsx}")
    if keep_crops:
        print(f"Kept generated crops: {crop_dir}")
    else:
        print("Removed temporary crops after embedding them in the workbook")


def count_reviewed_readings(workbook_path: Path) -> int:
    if not workbook_path.exists():
        return 0

    from openpyxl import load_workbook

    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )
    try:
        if "Meter Review" not in workbook.sheetnames:
            return 0
        sheet = workbook["Meter Review"]
        headers = {
            clean_text(cell.value): cell.column
            for cell in sheet[1]
        }
        reading_column = headers.get("Manual Reading")
        if reading_column is None:
            return 0
        return sum(
            valid_meter_reading(
                sheet.cell(row=row, column=reading_column).value
            )
            for row in range(2, sheet.max_row + 1)
        )
    finally:
        workbook.close()


def count_metadata_readings(metadata: pd.DataFrame) -> int:
    if "meter_reading" not in metadata.columns:
        return 0
    return int(metadata["meter_reading"].map(valid_meter_reading).sum())


def main() -> None:
    args = parse_args()

    if not args.photo_dir.exists():
        raise SystemExit(
            f"Photo directory does not exist: {args.photo_dir}"
        )

    heic_available = register_heic_support()

    metadata = load_metadata(
        metadata_csv=args.metadata_csv,
        photo_dir=args.photo_dir,
    )
    corrections = load_corrections(args.corrections_csv)
    metadata = apply_corrections(metadata, corrections)

    selected_photos = select_inventory_photos(metadata)

    if args.max_images is not None:
        selected_photos = selected_photos.head(
            args.max_images
        ).copy()

    if selected_photos.empty:
        raise SystemExit(
            "No matching photo files were found for the metadata CSV."
        )

    existing_readings = count_reviewed_readings(args.output_xlsx)
    replacement_readings = count_metadata_readings(selected_photos)
    if (
        replacement_readings < existing_readings
        and not args.allow_reading_loss
    ):
        raise SystemExit(
            "Refusing to replace the existing review workbook because the "
            f"new workbook would reduce six-digit readings from "
            f"{existing_readings} to {replacement_readings}. "
            "Correct the input inventory or use --allow-reading-loss only "
            "for deliberate recovery work."
        )

    build_workbook(
        selected_photos=selected_photos,
        output_xlsx=args.output_xlsx,
        crop_dir=args.crop_dir,
        heic_available=heic_available,
        keep_crops=args.keep_crops,
    )


if __name__ == "__main__":
    main()
