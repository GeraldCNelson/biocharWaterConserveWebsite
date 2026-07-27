#!/usr/bin/env python3
"""Tests for the safe, reproducible meter-photo review workbook builder."""

from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from PIL import Image


SCRIPT = (
    Path(__file__).parents[1]
    / "scripts"
    / "management"
    / "build_meter_review_workbook.py"
)
SPEC = importlib.util.spec_from_file_location(
    "build_meter_review_workbook_under_test",
    SCRIPT,
)
assert SPEC is not None and SPEC.loader is not None
MODULE = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(MODULE)


class MeterReviewWorkbookTests(unittest.TestCase):
    def test_load_metadata_preserves_leading_zero_reading(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            photo = root / "meter.jpg"
            Image.new("RGB", (100, 100), "white").save(photo)
            inventory = root / "inventory.csv"
            inventory.write_text(
                "filename,effective_datetime,sha256,meter_reading,"
                "review_status,notes\n"
                "meter.jpg,2026-07-24 08:25:17,"
                f"{'a' * 64},001234,readable,confirmed\n",
                encoding="utf-8",
            )

            metadata = MODULE.load_metadata(inventory, root)

            self.assertEqual(metadata.iloc[0]["meter_reading"], "001234")

    def test_correction_populates_new_photo(self) -> None:
        sha256 = "b" * 64
        metadata = pd.DataFrame(
            [
                {
                    "filename": "IMG_8600.jpg",
                    "sha256": sha256,
                    "meter_reading": "",
                    "review_status": "",
                    "notes": "",
                }
            ]
        )
        corrected = MODULE.apply_corrections(
            metadata,
            {
                sha256: {
                    "filename": "IMG_8600.jpg",
                    "meter_reading": "237972",
                    "review_status": "readable",
                    "notes": "User-confirmed start reading",
                }
            },
        )

        self.assertEqual(corrected.iloc[0]["meter_reading"], "237972")
        self.assertEqual(corrected.iloc[0]["review_status"], "readable")
        self.assertEqual(
            corrected.iloc[0]["notes"],
            "User-confirmed start reading",
        )

    def test_workbook_contains_preserved_fields_and_sha256(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            photo = root / "meter.jpg"
            Image.new("RGB", (640, 480), "white").save(photo)
            output = root / "review.xlsx"
            crop_dir = root / "crops"
            sha256 = "c" * 64
            selected = pd.DataFrame(
                [
                    {
                        "photo_datetime": pd.Timestamp(
                            "2026-07-24 08:25:17"
                        ),
                        "image_path": photo,
                        "meter_reading": "237972",
                        "review_status": "readable",
                        "notes": "Confirmed",
                        "sha256": sha256,
                        "duplicate_count": 1,
                    }
                ]
            )

            MODULE.build_workbook(
                selected_photos=selected,
                output_xlsx=output,
                crop_dir=crop_dir,
                heic_available=False,
                keep_crops=False,
            )

            workbook = load_workbook(output, data_only=True)
            sheet = workbook["Meter Review"]
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(
                headers,
                [
                    "Photo Date/Time",
                    "Photo",
                    "Manual Reading",
                    "Status",
                    "Notes",
                    "Filename",
                    "SHA-256",
                    "Crop Method",
                    "Duplicate Count",
                ],
            )
            self.assertEqual(str(sheet["C2"].value), "237972")
            self.assertEqual(sheet["D2"].value, "readable")
            self.assertEqual(sheet["E2"].value, "Confirmed")
            self.assertEqual(sheet["F2"].value, "meter.jpg")
            self.assertEqual(sheet["G2"].value, sha256)
            workbook.close()

    def test_counts_existing_six_digit_readings(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "review.xlsx"
            from openpyxl import Workbook

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Meter Review"
            sheet.append(["Manual Reading"])
            sheet.append(["237972"])
            sheet.append([""])
            sheet.append(["12345"])
            workbook.save(output)

            self.assertEqual(MODULE.count_reviewed_readings(output), 1)


if __name__ == "__main__":
    unittest.main()
