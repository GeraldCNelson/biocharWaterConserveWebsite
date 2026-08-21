#!/usr/bin/env python3
"""Build the dashboard field-biomass CSV from the master workbook.

The workbook sheets have changed layout over time.  This builder extracts only
field location, sampling date, and dry weight.  The reviewed 2023-2025 CSV is
kept as the historical authority; workbook values fill gaps and supply newer
sampling dates without rewriting established historical observations.
"""

from __future__ import annotations

import argparse
import os
import re
import tempfile
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from biochar_app.config.paths import (
    BIOCHAR_MASTER_WORKBOOK,
    BIOMASS_FIELD_CSV,
    BIOMASS_HISTORICAL_CSV,
)

BIOMASS_SHEETS = tuple(f"{year} BIOMASS" for year in range(2023, 2027))
LOCATION_RE = re.compile(r"^S[1-4][TMB]$")
DATE_RE = re.compile(r"(?P<month>\d{1,2})[/-](?P<day>\d{1,2})[/-](?P<year>\d{2,4})")


def _normalize_location(value: object) -> str | None:
    location = str(value).strip().upper() if value is not None else ""
    return location if LOCATION_RE.fullmatch(location) else None


def _parse_date(value: object) -> str | None:
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    match = DATE_RE.search(str(value)) if value is not None else None
    if not match:
        return None
    year = int(match.group("year"))
    if year < 100:
        year += 2000
    return date(year, int(match.group("month")), int(match.group("day"))).isoformat()


def extract_biomass_sheet(workbook_path: Path, sheet_name: str) -> pd.DataFrame:
    """Return location/sampling_date/dry_g rows from one biomass worksheet."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    location_header_row = next(
        i for i, row in enumerate(rows)
        if row and str(row[0]).strip().casefold() == "location"
    )
    date_row = rows[location_header_row - 1]
    header_row = rows[location_header_row]

    observations: list[dict[str, Any]] = []
    active_date: str | None = None
    dry_columns: list[tuple[int, str]] = []
    for column_index, header in enumerate(header_row):
        parsed_date = _parse_date(date_row[column_index] if column_index < len(date_row) else None)
        if parsed_date:
            active_date = parsed_date
        header_text = str(header).strip().casefold() if header is not None else ""
        if active_date and header_text in {"dry", "dry (g)"}:
            dry_columns.append((column_index, active_date))

    if not dry_columns:
        raise ValueError(f"No dated dry-weight columns found in worksheet {sheet_name!r}")

    for row in rows[location_header_row + 1:]:
        location = _normalize_location(row[0] if row else None)
        if location is None:
            continue
        for column_index, sampling_date in dry_columns:
            value = row[column_index] if column_index < len(row) else None
            dry_g = pd.to_numeric(value, errors="coerce")
            if pd.notna(dry_g):
                observations.append(
                    {"location": location, "sampling_date": sampling_date, "dry_g": float(dry_g)}
                )

    result = pd.DataFrame(observations, columns=["location", "sampling_date", "dry_g"])
    if result.duplicated(["location", "sampling_date"]).any():
        raise ValueError(f"Duplicate location/date observations in worksheet {sheet_name!r}")
    return result


def read_historical_biomass(path: Path) -> pd.DataFrame:
    """Normalize the reviewed two-header 2023-2025 CSV to wide form."""
    raw = pd.read_csv(path, header=[0, 1])
    location_column = raw.columns[0]
    normalized = pd.DataFrame({"location": raw[location_column].map(_normalize_location)})
    for column in raw.columns[1:]:
        sampling_date = _parse_date(column[0])
        if sampling_date:
            normalized[sampling_date] = pd.to_numeric(raw[column], errors="coerce")
    return normalized[normalized["location"].notna()].reset_index(drop=True)


def _observations_to_wide(observations: pd.DataFrame) -> pd.DataFrame:
    if observations.empty:
        return pd.DataFrame(columns=["location"])
    wide = observations.pivot(index="location", columns="sampling_date", values="dry_g")
    return wide.reset_index().rename_axis(columns=None)


def build_field_biomass(
    *,
    workbook_path: Path = BIOCHAR_MASTER_WORKBOOK,
    historical_path: Path = BIOMASS_HISTORICAL_CSV,
) -> pd.DataFrame:
    """Build wide biomass data, with reviewed historical values taking priority."""
    historical = read_historical_biomass(historical_path)
    extracted = pd.concat(
        [extract_biomass_sheet(workbook_path, sheet) for sheet in BIOMASS_SHEETS],
        ignore_index=True,
    )
    workbook_wide = _observations_to_wide(extracted)
    historical_dates = [column for column in historical.columns if column != "location"]
    historical_latest = max(historical_dates)
    allowed_workbook_dates = [
        column for column in workbook_wide.columns
        if column == "location" or column in historical_dates or column > historical_latest
    ]
    workbook_wide = workbook_wide[allowed_workbook_dates]
    combined = historical.set_index("location").combine_first(workbook_wide.set_index("location"))
    combined = combined.reindex(sorted(combined.columns), axis=1)
    location_order = [f"S{strip}{position}" for strip in range(1, 5) for position in "TMB"]
    combined = combined.reindex(location_order).round(10).reset_index()
    expected_locations = {f"S{strip}{position}" for strip in range(1, 5) for position in "TMB"}
    if set(combined["location"]) != expected_locations:
        raise ValueError("Biomass output must contain exactly the 12 expected field locations")
    return combined


def _atomic_write_csv(dataframe: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", dir=path.parent, delete=False) as handle:
        temporary_path = Path(handle.name)
        dataframe.to_csv(handle, index=False)
    try:
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def build_and_install_field_biomass(
    *,
    workbook_path: Path = BIOCHAR_MASTER_WORKBOOK,
    historical_path: Path = BIOMASS_HISTORICAL_CSV,
    output_path: Path = BIOMASS_FIELD_CSV,
) -> dict[str, Any]:
    data = build_field_biomass(workbook_path=workbook_path, historical_path=historical_path)
    _atomic_write_csv(data, output_path)
    return {
        "built_at_utc": datetime.now(timezone.utc).isoformat(),
        "rows": len(data),
        "sampling_dates": len(data.columns) - 1,
        "latest_sampling_date": data.columns[-1],
        "output_path": str(output_path),
    }


def main() -> None:
    """Compatibility entry point; ETL is now the preferred orchestrator."""
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=BIOCHAR_MASTER_WORKBOOK)
    args = parser.parse_args()
    print(build_and_install_field_biomass(workbook_path=args.workbook))


if __name__ == "__main__":
    # TODO(delete after ETL adoption): remove this standalone entry point once
    # operations have used biochar_app/scripts/etl.py for a full update cycle.
    main()
