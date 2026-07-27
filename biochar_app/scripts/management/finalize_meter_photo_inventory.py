#!/usr/bin/env python3
"""Merge reviewed meter readings and build the unique photo inventory.

Script metadata
---------------
Purpose:
    Reproducibly convert reviewed meter-photo workbooks plus the canonical
    metadata inventory into ``photo_inventory_unique.csv``.
Primary inputs:
    ``photo_inventory.csv``, the main meter-review workbook, the
    full-resolution follow-up workbook, and an optional corrections ledger.
Primary outputs:
    Unique inventory CSV, keep/exclude manifest CSV, and audit JSON.
Authority order:
    Explicit correction -> readable exact main-workbook entry -> unambiguous
    readable main-workbook photo-family consensus -> full-resolution follow-up
    -> existing inventory value.
Safety:
    Input files are read-only. The optional unique-photo directory must not
    already exist, preventing accidental mixing with an earlier run.
Introduced:
    2026-07-23, following recovery and review of irrigation meter photographs.
Maintainer:
    Biochar Water Conservation project (Gerald Nelson).
Version:
    0.1.0
Example:
    Run from the repository root (replace workbook names or paths if needed)::

        python biochar_app/scripts/management/finalize_meter_photo_inventory.py \
          --inventory-csv biochar_app/data-processed/management/irrigation/photos/photo_inventory.csv \
          --main-workbook biochar_app/data-processed/management/irrigation/photos/meter_photo_review_updated_2026-07-23.xlsx \
          --full-resolution-workbook biochar_app/data-processed/management/irrigation/photos/meter_photo_unreadable_full_resolution_review_updated_2026-07-23.xlsx \
          --corrections-csv biochar_app/data-processed/management/irrigation/photos/meter_photo_reading_corrections.csv \
          --output-csv biochar_app/data-processed/management/irrigation/photos/photo_inventory_unique.csv \
          --manifest-csv biochar_app/data-processed/management/irrigation/photos/photo_inventory_unique_manifest.csv \
          --audit-json biochar_app/data-processed/management/irrigation/photos/photo_inventory_unique_audit.json

This is the reproducible bridge between:

* photo_inventory.csv (metadata for every archived photo),
* the main meter-review workbook,
* the full-resolution follow-up workbook, and
* an explicit correction ledger.

The script never edits its inputs. It writes a unique inventory, a complete
selection manifest, and a machine-readable audit report.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


__version__ = "0.1.0"

EXPECTED_READING_DIGITS = 6
READABLE_STATUSES = {"readable"}
RENDITION_SUFFIX = re.compile(r"_(?:1_102_o|1_102_a|1_105_c|4_5005_c)$", re.I)


@dataclass(frozen=True)
class ReviewReading:
    sha256: str
    filename: str
    reading: str
    status: str
    notes: str
    source: str


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "<na>"}:
        return ""
    return text


def normalize_reading(value: object) -> str:
    text = clean(value)
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return text


def valid_reading(value: object) -> bool:
    return bool(re.fullmatch(rf"\d{{{EXPECTED_READING_DIGITS}}}", normalize_reading(value)))


def photo_family(filename: str) -> str:
    stem = Path(clean(filename)).stem
    return RENDITION_SUFFIX.sub("", stem).casefold()


def integer(value: object) -> int:
    try:
        return int(float(clean(value)))
    except (TypeError, ValueError):
        return 0


def read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), [dict(row) for row in reader]


def write_csv_rows(path: Path, fieldnames: list[str], rows: Iterable[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def workbook_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"{path.name} does not contain sheet {sheet_name!r}")
    sheet = workbook[sheet_name]
    values = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(values)]
    return [dict(zip(headers, row)) for row in values]


def read_main_review(path: Path) -> list[ReviewReading]:
    rows = workbook_rows(path, "Meter Review")
    required = {"Filename", "SHA-256", "Manual Reading", "Status", "Notes"}
    missing = required.difference(rows[0] if rows else {})
    if missing:
        raise KeyError(f"Main review workbook is missing columns: {sorted(missing)}")
    return [
        ReviewReading(
            sha256=clean(row["SHA-256"]),
            filename=clean(row["Filename"]),
            reading=normalize_reading(row["Manual Reading"]),
            status=clean(row["Status"]).casefold(),
            notes=clean(row["Notes"]),
            source="main_review",
        )
        for row in rows
        if clean(row["SHA-256"])
    ]


def read_full_resolution_review(path: Path) -> list[ReviewReading]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    rows: list[ReviewReading] = []
    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith("Photo "):
            continue
        sheet = workbook[sheet_name]
        sha256 = clean(sheet["B6"].value)
        if not sha256:
            continue
        rows.append(
            ReviewReading(
                sha256=sha256,
                filename=clean(sheet["B5"].value),
                reading=normalize_reading(sheet["B11"].value),
                status=clean(sheet["B12"].value).casefold(),
                notes=clean(sheet["B13"].value),
                source="full_resolution_followup",
            )
        )
    return rows


def read_corrections(path: Path | None) -> dict[str, ReviewReading]:
    if path is None or not path.exists():
        return {}
    _, rows = read_csv_rows(path)
    required = {"sha256", "filename", "corrected_reading", "reason"}
    missing = required.difference(rows[0] if rows else {})
    if missing:
        raise KeyError(f"Correction ledger is missing columns: {sorted(missing)}")
    result: dict[str, ReviewReading] = {}
    for row in rows:
        reading = normalize_reading(row["corrected_reading"])
        if not valid_reading(reading):
            raise ValueError(f"Invalid correction reading for {row['filename']}: {reading!r}")
        result[clean(row["sha256"])] = ReviewReading(
            sha256=clean(row["sha256"]),
            filename=clean(row["filename"]),
            reading=reading,
            status="readable",
            notes=clean(row["reason"]),
            source="explicit_correction",
        )
    return result


def main_family_consensus(
    main_rows: list[ReviewReading],
) -> tuple[dict[str, str], dict[str, list[str]]]:
    readings: dict[str, set[str]] = defaultdict(set)
    for row in main_rows:
        if row.status in READABLE_STATUSES and valid_reading(row.reading):
            readings[photo_family(row.filename)].add(row.reading)
    ambiguous = {
        family: sorted(values)
        for family, values in readings.items()
        if len(values) > 1
    }
    consensus = {
        family: next(iter(values))
        for family, values in readings.items()
        if len(values) == 1
    }
    return consensus, ambiguous


def choose_reading(
    inventory_row: dict[str, str],
    main_by_sha: dict[str, ReviewReading],
    followup_by_sha: dict[str, ReviewReading],
    corrections: dict[str, ReviewReading],
    family_consensus: dict[str, str],
) -> tuple[str, str, str, list[dict[str, str]]]:
    sha256 = clean(inventory_row.get("sha256"))
    filename = clean(inventory_row.get("filename"))
    family = photo_family(filename)
    main_row = main_by_sha.get(sha256)
    followup_row = followup_by_sha.get(sha256)
    correction = corrections.get(sha256)
    conflicts: list[dict[str, str]] = []

    if correction is not None:
        return correction.reading, "readable", correction.source, conflicts

    if main_row and main_row.status in READABLE_STATUSES and valid_reading(main_row.reading):
        selected = main_row
    elif family in family_consensus:
        selected = ReviewReading(
            sha256=sha256,
            filename=filename,
            reading=family_consensus[family],
            status="readable",
            notes="Propagated from readable main-workbook rendition",
            source="main_review_family_consensus",
        )
    elif followup_row and valid_reading(followup_row.reading):
        selected = followup_row
    elif main_row and valid_reading(main_row.reading):
        selected = main_row
    else:
        existing = normalize_reading(inventory_row.get("meter_reading"))
        status = clean(inventory_row.get("review_status")).casefold()
        return existing, status, "existing_inventory", conflicts

    if (
        followup_row
        and valid_reading(followup_row.reading)
        and followup_row.reading != selected.reading
        and selected.source != "full_resolution_followup"
    ):
        conflicts.append(
            {
                "sha256": sha256,
                "filename": filename,
                "selected_reading": selected.reading,
                "selected_source": selected.source,
                "suppressed_followup_reading": followup_row.reading,
            }
        )
    return selected.reading, selected.status or "readable", selected.source, conflicts


def preferred_record(rows: list[dict[str, str]]) -> dict[str, str]:
    return sorted(
        rows,
        key=lambda row: (
            -(integer(row.get("image_width")) * integer(row.get("image_height"))),
            -integer(row.get("file_size_bytes")),
            clean(row.get("filename")).casefold(),
        ),
    )[0]


def finalize_inventory(
    inventory_csv: Path,
    main_workbook: Path,
    full_resolution_workbook: Path,
    corrections_csv: Path | None,
    output_csv: Path,
    manifest_csv: Path,
    audit_json: Path,
    copy_unique_dir: Path | None = None,
    originals_dir: Path | None = None,
) -> dict[str, object]:
    fieldnames, inventory = read_csv_rows(inventory_csv)
    required = {
        "filename", "sha256", "effective_datetime", "meter_reading",
        "review_status", "image_width", "image_height", "file_size_bytes",
    }
    missing = required.difference(fieldnames)
    if missing:
        raise KeyError(f"Inventory is missing columns: {sorted(missing)}")

    main_rows = read_main_review(main_workbook)
    followup_rows = read_full_resolution_review(full_resolution_workbook)
    main_by_sha = {row.sha256: row for row in main_rows}
    followup_by_sha = {row.sha256: row for row in followup_rows}
    corrections = read_corrections(corrections_csv)
    family_consensus, ambiguous_main_families = main_family_consensus(main_rows)

    source_counts: dict[str, int] = defaultdict(int)
    suppressed_conflicts: list[dict[str, str]] = []
    merged: list[dict[str, str]] = []
    for source_row in inventory:
        row = dict(source_row)
        reading, status, source, conflicts = choose_reading(
            row, main_by_sha, followup_by_sha, corrections, family_consensus
        )
        row["meter_reading"] = reading
        row["review_status"] = status
        row["meter_reading_source"] = source
        if "meter_reading_source" not in fieldnames:
            fieldnames.append("meter_reading_source")
        source_counts[source] += 1
        suppressed_conflicts.extend(conflicts)
        merged.append(row)

    duplicate_groups: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in merged:
        timestamp = clean(row.get("effective_datetime"))
        reading = normalize_reading(row.get("meter_reading"))
        if timestamp and valid_reading(reading):
            duplicate_groups[(timestamp, reading)].append(row)

    selected_shas: set[str] = set()
    group_details: dict[str, dict[str, str]] = {}
    group_number = 0
    for key, rows in sorted(duplicate_groups.items()):
        if len(rows) == 1:
            selected_shas.add(clean(rows[0]["sha256"]))
            continue
        group_number += 1
        selected = preferred_record(rows)
        selected_sha = clean(selected["sha256"])
        selected_shas.add(selected_sha)
        group_id = f"DUP-{group_number:03d}"
        for row in rows:
            group_details[clean(row["sha256"])] = {
                "duplicate_group": group_id,
                "selected_filename": clean(selected["filename"]),
                "action": "kept" if clean(row["sha256"]) == selected_sha else "excluded_from_unique",
            }

    for row in merged:
        reading = normalize_reading(row.get("meter_reading"))
        timestamp = clean(row.get("effective_datetime"))
        if not (timestamp and valid_reading(reading)):
            selected_shas.add(clean(row["sha256"]))

    selected = [row for row in merged if clean(row["sha256"]) in selected_shas]
    manifest_fields = [
        "action", "duplicate_group", "selected_filename",
        "effective_datetime", "meter_reading", "meter_reading_source",
        "filename", "sha256", "image_width", "image_height", "file_size_bytes",
    ]
    manifest = []
    for row in merged:
        detail = group_details.get(clean(row["sha256"]), {})
        manifest.append(
            {
                "action": detail.get("action", "kept"),
                "duplicate_group": detail.get("duplicate_group", ""),
                "selected_filename": detail.get("selected_filename", clean(row["filename"])),
                **row,
            }
        )

    write_csv_rows(output_csv, fieldnames, selected)
    write_csv_rows(manifest_csv, manifest_fields, manifest)

    copied = 0
    if copy_unique_dir is not None:
        if originals_dir is None:
            raise ValueError("--copy-unique-dir requires --originals-dir")
        copy_unique_dir.mkdir(parents=True, exist_ok=False)
        for row in selected:
            shutil.copy2(originals_dir / row["filename"], copy_unique_dir / row["filename"])
            copied += 1

    audit = {
        "inputs": {
            "inventory_csv": str(inventory_csv),
            "main_workbook": str(main_workbook),
            "full_resolution_workbook": str(full_resolution_workbook),
            "corrections_csv": str(corrections_csv) if corrections_csv else None,
        },
        "outputs": {
            "unique_inventory_csv": str(output_csv),
            "manifest_csv": str(manifest_csv),
            "copy_unique_dir": str(copy_unique_dir) if copy_unique_dir else None,
        },
        "inventory_rows": len(inventory),
        "selected_unique_rows": len(selected),
        "excluded_duplicate_rows": len(inventory) - len(selected),
        "duplicate_groups": sum(1 for rows in duplicate_groups.values() if len(rows) > 1),
        "six_digit_unique_readings": sum(valid_reading(row["meter_reading"]) for row in selected),
        "blank_unique_readings": sum(not clean(row["meter_reading"]) for row in selected),
        "reading_source_counts": dict(sorted(source_counts.items())),
        "explicit_correction_count": len(corrections),
        "ambiguous_main_workbook_families": ambiguous_main_families,
        "suppressed_followup_conflicts": suppressed_conflicts,
        "copied_unique_files": copied,
        "output_sha256": hashlib.sha256(output_csv.read_bytes()).hexdigest(),
    }
    audit_json.parent.mkdir(parents=True, exist_ok=True)
    audit_json.write_text(json.dumps(audit, indent=2), encoding="utf-8")
    return audit


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--inventory-csv", type=Path, required=True)
    parser.add_argument("--main-workbook", type=Path, required=True)
    parser.add_argument("--full-resolution-workbook", type=Path, required=True)
    parser.add_argument("--corrections-csv", type=Path)
    parser.add_argument("--output-csv", type=Path, required=True)
    parser.add_argument("--manifest-csv", type=Path, required=True)
    parser.add_argument("--audit-json", type=Path, required=True)
    parser.add_argument("--copy-unique-dir", type=Path)
    parser.add_argument("--originals-dir", type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    audit = finalize_inventory(
        inventory_csv=args.inventory_csv,
        main_workbook=args.main_workbook,
        full_resolution_workbook=args.full_resolution_workbook,
        corrections_csv=args.corrections_csv,
        output_csv=args.output_csv,
        manifest_csv=args.manifest_csv,
        audit_json=args.audit_json,
        copy_unique_dir=args.copy_unique_dir,
        originals_dir=args.originals_dir,
    )
    print(json.dumps(audit, indent=2))


if __name__ == "__main__":
    main()
